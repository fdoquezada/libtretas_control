from django.contrib import admin
from django.http import HttpResponse
import openpyxl
from datetime import datetime
from .models import Conductor, RegistroSemanal

@admin.register(Conductor)
class ConductorAdmin(admin.ModelAdmin):
    list_display = ('nombre', 'rut')
    search_fields = ('nombre', 'rut')
    list_per_page = 20
    ordering = ('nombre',)
    actions = ['exportar_a_excel']

    def exportar_a_excel(self, request, queryset):
        # Crear un nuevo libro de Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Conductores"

        # Agregar encabezados
        ws['A1'] = 'Nombre'
        ws['B1'] = 'RUT'

        # Agregar datos
        for idx, conductor in enumerate(queryset, start=2):
            ws[f'A{idx}'] = conductor.nombre
            ws[f'B{idx}'] = conductor.rut

        # Configurar el nombre del archivo
        fecha_actual = datetime.now().strftime('%Y-%m-%d')
        filename = f'conductores_{fecha_actual}.xlsx'

        # Crear la respuesta HTTP con el archivo Excel
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        # Guardar el libro de Excel en la respuesta
        wb.save(response)
        return response

    exportar_a_excel.short_description = "Exportar seleccionados a Excel"

@admin.register(RegistroSemanal)
class RegistroSemanalAdmin(admin.ModelAdmin):
    # ... existing code ...
    actions = ['marcar_como_entregado', 'marcar_como_no_entregado', 'exportar_a_excel']

    def exportar_a_excel(self, request, queryset):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Registros Semanales"

        # Encabezados
        headers = [
            'Conductor', 'RUT', 'Semana', 'Año', 
            'Entregado', 'Fecha de Entrega'
        ]
        for col, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col, value=header)

        # Datos
        for idx, registro in enumerate(queryset, start=2):
            ws.cell(row=idx, column=1, value=registro.conductor.nombre)
            ws.cell(row=idx, column=2, value=registro.conductor.rut)
            ws.cell(row=idx, column=3, value=registro.semana)
            ws.cell(row=idx, column=4, value=registro.año)
            ws.cell(row=idx, column=5, value='Sí' if registro.entregado else 'No')
            ws.cell(row=idx, column=6, value=registro.fecha_entrega.strftime('%d/%m/%Y') if registro.fecha_entrega else '')

        # Ajustar ancho de columnas
        for column in ws.columns:
            max_length = 0
            column_letter = openpyxl.utils.get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Configurar respuesta
        fecha_actual = datetime.now().strftime('%Y-%m-%d')
        filename = f'registros_semanales_{fecha_actual}.xlsx'
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response

    exportar_a_excel.short_description = "Exportar seleccionados a Excel"
